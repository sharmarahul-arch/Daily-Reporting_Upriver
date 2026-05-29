"""
Interactive capture of Amazon Seller Central account IDs.

For each active account, this script opens the SC account switcher in a
headed browser, prompts the user to manually click through to the target
account, then captures `mons_sel_dir_mcid` / `mons_sel_mkid` /
`mons_sel_dir_paid` from the resulting URL and saves them to accounts.xlsx.

Once all accounts have IDs, switch_sc_account() uses direct-URL navigation
to switch sessions, bypassing the unreliable switcher UI clicker.
"""

import asyncio
import logging
import re
from typing import Optional

import openpyxl
from playwright.async_api import async_playwright

from src.config import (
    CONFIG_DIR, load_accounts, SC_BASE_URL, SC_BASE_URL_US,
    AMAZON_EMAIL, AMAZON_PASSWORD,
)
from src.auth import (
    restore_session, save_session, login_seller_central,
)

log = logging.getLogger(__name__)


_ID_PATTERNS = {
    "mcid": re.compile(r"mons_sel_dir_mcid=([^&]+)"),
    "mkid": re.compile(r"mons_sel_mkid=([^&]+)"),
    "paid": re.compile(r"mons_sel_dir_paid=([^&]+)"),
}


def _extract_ids(url: str) -> dict:
    out = {}
    for k, p in _ID_PATTERNS.items():
        m = p.search(url)
        if m:
            out[k] = m.group(1)
    return out


def _load_existing_ids() -> dict:
    """Return {account_name: {mcid, mkid, paid}} from accounts.xlsx."""
    xlsx = CONFIG_DIR / "accounts.xlsx"
    if not xlsx.exists():
        return {}
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Accounts"]
    hdr_row = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip() == "Account Name":
                hdr_row = cell.row
                break
        if hdr_row:
            break
    if not hdr_row:
        return {}
    headers = {}
    for cell in ws[hdr_row]:
        v = str(cell.value or "").strip()
        if v:
            headers[v] = cell.column
    if "SC Merchant ID" not in headers:
        return {}
    name_col = headers["Account Name"]
    out = {}
    for row in ws.iter_rows(min_row=hdr_row + 2, values_only=False):
        n = str(row[name_col - 1].value or "").strip()
        if not n:
            continue
        mcid = str(row[headers["SC Merchant ID"] - 1].value or "").strip()
        mkid = str(row[headers.get("SC Marketplace ID", 0) - 1].value or "").strip() if "SC Marketplace ID" in headers else ""
        paid = str(row[headers.get("SC Paid ID", 0) - 1].value or "").strip() if "SC Paid ID" in headers else ""
        if mcid:
            out[n] = {"mcid": mcid, "mkid": mkid, "paid": paid}
    return out


def _save_ids(captured: dict):
    """Write captured IDs into accounts.xlsx (adds columns if missing)."""
    xlsx = CONFIG_DIR / "accounts.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Accounts"]
    hdr_row = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip() == "Account Name":
                hdr_row = cell.row
                break
        if hdr_row:
            break

    headers = {}
    last_col = 0
    for cell in ws[hdr_row]:
        v = str(cell.value or "").strip()
        if v:
            headers[v] = cell.column
            last_col = max(last_col, cell.column)

    for col_name in ("SC Merchant ID", "SC Marketplace ID", "SC Paid ID"):
        if col_name not in headers:
            last_col += 1
            ws.cell(row=hdr_row, column=last_col, value=col_name)
            headers[col_name] = last_col

    name_col = headers["Account Name"]
    n_written = 0
    for row in ws.iter_rows(min_row=hdr_row + 2):
        n = str(row[name_col - 1].value or "").strip()
        if n in captured:
            ids = captured[n]
            row[headers["SC Merchant ID"] - 1].value = ids.get("mcid", "")
            row[headers["SC Marketplace ID"] - 1].value = ids.get("mkid", "")
            row[headers["SC Paid ID"] - 1].value = ids.get("paid", "")
            n_written += 1
    wb.save(xlsx)
    return n_written


async def _capture_for_account(page, account: dict, switcher_url: str) -> Optional[dict]:
    """Open the switcher, prompt user to click through, then poll URL for new IDs."""
    name = account["name"]
    target = account.get("sc_account_name", name)
    marketplace = account.get("marketplace", "IN")
    print()
    print("=" * 64)
    print(f"  Account: {name}  ({marketplace})")
    print(f"  → In the browser, switch to:  {target}")
    print("=" * 64)
    input("  Press ENTER to open the switcher → ")

    await page.goto(switcher_url, wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(1500)

    initial_url = page.url
    initial_ids = _extract_ids(initial_url)
    print(f"  Switcher is open. Now click '{target}' → pick its marketplace → 'Select account'.")
    print("  (Polling URL for the new IDs — you have up to 5 minutes.)")

    deadline = asyncio.get_event_loop().time() + 300
    new_ids = None
    while asyncio.get_event_loop().time() < deadline:
        await asyncio.sleep(1.2)
        cur = page.url
        ids = _extract_ids(cur)
        # We want a URL that has all three IDs AND is different from the initial
        if ids.get("mcid") and ids != initial_ids and "homepage" in cur:
            new_ids = ids
            break

    if not new_ids:
        print(f"  ⚠️  Timed out. Skipping {name}.")
        return None

    print(f"  ✓ Captured")
    print(f"     mcid = {new_ids.get('mcid','')[:50]}")
    print(f"     mkid = {new_ids.get('mkid','')[:50]}")
    print(f"     paid = {new_ids.get('paid','')[:50]}")
    return new_ids


async def capture_all_ids(force: bool = False):
    """
    Walk through all active accounts and capture SC IDs interactively.
    If force=False (default), skip accounts that already have IDs in accounts.xlsx.
    """
    accounts = load_accounts()
    existing = _load_existing_ids()

    to_capture = [a for a in accounts if force or a["name"] not in existing]
    if not to_capture:
        print(f"All {len(accounts)} accounts already have SC IDs. Use force=True to re-capture.")
        return

    print(f"\nAccounts already captured: {sorted(existing.keys())}")
    print(f"Accounts to capture this run ({len(to_capture)}): "
          + ", ".join(a["name"] for a in to_capture))
    print()

    has_us = any(a.get("marketplace", "IN").upper() == "US" for a in to_capture)
    has_in = any(a.get("marketplace", "IN").upper() != "US" for a in to_capture)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False, args=["--no-sandbox"])

        in_ctx = in_page = None
        if has_in:
            in_ctx = await browser.new_context(viewport={"width": 1500, "height": 900})
            in_page = await in_ctx.new_page()
            if not await restore_session(in_ctx, "sc"):
                ok = await login_seller_central(in_page, AMAZON_EMAIL, AMAZON_PASSWORD, base_url=SC_BASE_URL)
                if ok:
                    await save_session(in_ctx, "sc")

        us_ctx = us_page = None
        if has_us:
            us_ctx = await browser.new_context(viewport={"width": 1500, "height": 900})
            us_page = await us_ctx.new_page()
            if not await restore_session(us_ctx, "sc_us"):
                ok = await login_seller_central(us_page, AMAZON_EMAIL, AMAZON_PASSWORD, base_url=SC_BASE_URL_US)
                if ok:
                    await save_session(us_ctx, "sc_us")

        captured = {}
        for acct in to_capture:
            is_us = acct.get("marketplace", "IN").upper() == "US"
            page = us_page if is_us else in_page
            home = SC_BASE_URL_US if is_us else SC_BASE_URL
            switcher_url = f"{home}/account-switcher/default/merchantMarketplace?returnTo=%2Fgp%2Fhomepage.html"
            try:
                ids = await _capture_for_account(page, acct, switcher_url)
                if ids:
                    captured[acct["name"]] = ids
            except Exception as exc:
                print(f"  Error capturing {acct['name']}: {exc}")

        await browser.close()

    if captured:
        n = _save_ids(captured)
        print(f"\n✓ Wrote {n} accounts to config/accounts.xlsx")
    else:
        print("\nNo IDs captured.")
