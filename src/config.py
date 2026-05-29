import json
from pathlib import Path
from typing import List

BASE_DIR = Path(__file__).parent.parent
CONFIG_DIR = BASE_DIR / "config"
SESSIONS_DIR = BASE_DIR / "sessions"
DOWNLOADS_DIR = BASE_DIR / "downloads"
LOGS_DIR = BASE_DIR / "logs"


def _load_credentials() -> dict:
    path = CONFIG_DIR / "credentials.json"
    if not path.exists():
        raise FileNotFoundError(
            f"credentials.json not found at {path}.\n"
            "Copy config/credentials.json, fill in your Amazon login and "
            "Google service account path, then run again."
        )
    with open(path) as f:
        return json.load(f)


_creds = _load_credentials()

AMAZON_EMAIL    = _creds["amazon_email"]
AMAZON_PASSWORD = _creds["amazon_password"]
SHEETS_CREDS_PATH = _creds["sheets_service_account_path"]

# Chrome User Data directory (macOS default).
# IMPORTANT: Close Google Chrome before running the script.
CHROME_USER_DATA_DIR = Path.home() / "Library" / "Application Support" / "Google" / "Chrome"

SC_BASE_URL      = "https://sellercentral.amazon.in"
SC_BASE_URL_US   = "https://sellercentral.amazon.com"
SC_SWITCHER_URL  = "https://sellercentral.amazon.in/account-switcher/switcher"

ADS_BASE_URL            = "https://advertising.amazon.in"
ADS_CAMPAIGN_MANAGER    = "https://advertising.amazon.in/campaign-manager"

ADS_BASE_URL_US         = "https://advertising.amazon.com"
ADS_CAMPAIGN_MANAGER_US = "https://advertising.amazon.com/campaign-manager"

# Google Sheet tab names (report_key → tab name)
SHEET_TABS = {
    "sales":               "Sales",
    "inventory":           "Inventory",
    "campaigns":           "Ad Campaigns",
    "advertised_products": "Advertised Products",
}


def load_accounts() -> List[dict]:
    """
    Load accounts from config/accounts.xlsx (the human-editable Excel file).
    Falls back to accounts.json if the xlsx is not present.
    """
    xlsx_path = CONFIG_DIR / "accounts.xlsx"
    json_path = CONFIG_DIR / "accounts.json"

    if xlsx_path.exists():
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["Accounts"]

        # Find header row (the row containing "Account Name")
        header_row = None
        for row in ws.iter_rows():
            for cell in row:
                if str(cell.value or "").strip() == "Account Name":
                    header_row = cell.row
                    break
            if header_row:
                break

        if not header_row:
            raise ValueError("Could not find header row in accounts.xlsx — make sure 'Account Name' column exists.")

        # Read column positions from header row
        headers = {}
        for cell in ws[header_row]:
            val = str(cell.value or "").strip()
            if val:
                headers[val] = cell.column - 1  # 0-based index

        accounts = []
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):  # +2 skips tooltip row
            # Skip blank rows and template placeholder rows
            name = str(row[headers.get("Account Name", 0)] or "").strip()
            if not name or name.lower() in ("account name", "new brand name"):
                continue

            def col(key, default=""):
                idx = headers.get(key)
                return str(row[idx] or "").strip() if idx is not None else default

            active_val = col("Active", "Yes").lower()
            if active_val not in ("yes", "y", "true", "1"):
                continue

            accounts.append({
                "name":             name,
                "sc_account_name":  col("SC Account Name", name),
                "sc_parent":        col("SC Parent") or None,
                "ads_account_name": col("Ads Account Name", name),
                "marketplace":      col("Marketplace", "IN").upper(),
                "google_sheet_id":  col("Google Sheet ID"),
                "email":            col("Amazon Email") or AMAZON_EMAIL,
                "password":         col("Amazon Password") or AMAZON_PASSWORD,
                # Pre-captured SC IDs (from --capture-ids). When set, the
                # switcher bypasses Amazon's UI and navigates directly.
                "sc_merchant_id":   col("SC Merchant ID") or None,
                "sc_marketplace_id":col("SC Marketplace ID") or None,
                "sc_paid_id":       col("SC Paid ID") or None,
                "sc_dc":            col("SC DC") or None,
                "active":           True,
            })
        return accounts

    # Fallback: legacy accounts.json
    with open(json_path) as f:
        accounts = json.load(f)
    return [a for a in accounts if a.get("active", True)]

